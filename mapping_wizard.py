import tkinter as tk
from tkinter import messagebox, filedialog
import json
from openpyxl import load_workbook


# ================== HELPERS ==================

def find_all_cells(sheet, keyword):
    positions = []
    if not keyword:
        return positions

    for r in range(sheet.max_row):
        for c in range(sheet.max_column):
            val = sheet.cell(row=r + 1, column=c + 1).value
            if val and keyword in str(val):
                positions.append((r, c))
    return positions


def extract_block(sheet, start_row, start_col, stop_col):
    data = []
    r = start_row
    

    while r < sheet.max_row:
        row_data = []
        for c in range(start_col, stop_col + 1):
            row_data.append(sheet.cell(row=r + 1, column=c + 1).value)

        if all(v is None for v in row_data):
            break

        data.append(row_data)
        r += 1

    return data


# ================== APP ==================

class MappingWizardApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generic Excel Mapping Wizard + Looping")
        self.root.geometry("950x680")

        self.rules = []
        self.excel_file = tk.StringVar()

        # ===== Excel File =====
        tk.Label(root, text="Excel File").grid(row=0, column=0, sticky="w")
        tk.Entry(root, textvariable=self.excel_file, width=60).grid(row=0, column=1, sticky="w")
        tk.Button(root, text="Browse", command=self.browse_excel).grid(row=0, column=2)

        # ===== Rule Form =====
        tk.Label(root, text="FROM Keyword").grid(row=1, column=0, sticky="w")
        self.from_entry = tk.Entry(root, width=40)
        self.from_entry.grid(row=1, column=1, sticky="w")

        tk.Label(root, text="Skip Rows").grid(row=2, column=0, sticky="w")
        self.skip_rows = tk.Entry(root, width=5)
        self.skip_rows.grid(row=2, column=1, sticky="w")

        tk.Label(root, text="Skip Cols").grid(row=3, column=0, sticky="w")
        self.skip_cols = tk.Entry(root, width=5)
        self.skip_cols.grid(row=3, column=1, sticky="w")

        tk.Label(root, text="Extract Start (rows, cols)").grid(row=4, column=0, sticky="w")
        self.start_rows = tk.Entry(root, width=5)
        self.start_cols = tk.Entry(root, width=5)
        self.start_rows.grid(row=4, column=1, sticky="w")
        self.start_cols.grid(row=4, column=1, padx=60, sticky="w")

        tk.Label(root, text="Extract Stop Col").grid(row=5, column=0, sticky="w")
        self.stop_cols = tk.Entry(root, width=5)
        self.stop_cols.grid(row=5, column=1, sticky="w")

        # ===== LOOP CHECKBOX =====
        self.loop_var = tk.BooleanVar()
        tk.Checkbutton(root, text="Loop (repeat for every occurrence)", variable=self.loop_var)\
            .grid(row=6, column=1, sticky="w")

        tk.Label(root, text="SECTION Name (AS)").grid(row=7, column=0, sticky="w")
        self.section_entry = tk.Entry(root, width=40)
        self.section_entry.grid(row=7, column=1, sticky="w")

        tk.Button(root, text="Add Rule", command=self.add_rule).grid(row=8, column=1, pady=10, sticky="w")

        # ===== Rule List =====
        self.listbox = tk.Listbox(root, width=120, height=10)
        self.listbox.grid(row=9, column=0, columnspan=3, padx=10, pady=10)

        # ===== Buttons =====
        tk.Button(root, text="Export Rules JSON", command=self.export_json).grid(row=10, column=0)
        tk.Button(root, text="Run Extraction", command=self.run_extraction).grid(row=10, column=1)

    # ================== METHODS ==================

    def browse_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if path:
            self.excel_file.set(path)

    def safe_int(self, val):
        return int(val) if val.strip() else 0

    def add_rule(self):
        if not self.excel_file.get():
            messagebox.showerror("Error", "Select Excel file first")
            return

        if not self.section_entry.get().strip():
            messagebox.showerror("Error", "SECTION name required")
            return

        rule = {
            "from": self.from_entry.get().strip(),
            "skip": {
                "rows": self.safe_int(self.skip_rows.get()),
                "cols": self.safe_int(self.skip_cols.get())
            },
            "extract": {
                "start": {
                    "rows": self.safe_int(self.start_rows.get()),
                    "cols": self.safe_int(self.start_cols.get())
                },
                "stop_col": self.safe_int(self.stop_cols.get())
            },
            "loop": self.loop_var.get(),
            "as": self.section_entry.get().strip()
        }

        self.rules.append(rule)
        self.listbox.insert(tk.END, json.dumps(rule))
        self.clear_fields()

    def clear_fields(self):
        for e in [
            self.from_entry, self.skip_rows, self.skip_cols,
            self.start_rows, self.start_cols,
            self.stop_cols, self.section_entry
        ]:
            e.delete(0, tk.END)
        self.loop_var.set(False)

    def export_json(self):
        with open("mapping_rules.json", "w") as f:
            json.dump({
                "excel_file": self.excel_file.get(),
                "rules": self.rules
            }, f, indent=4)
        messagebox.showinfo("Saved", "mapping_rules.json created")

    def run_extraction(self):
        wb = load_workbook(self.excel_file.get(), data_only=True)
        sheet = wb.active

        extracted = {}

        for rule in self.rules:
            positions = find_all_cells(sheet, rule["from"])
            if not positions:
                continue

            if not rule["loop"]:
                positions = positions[:1]

            extracted[rule["as"]] = []

            for ar, ac in positions:
                base_row = ar + rule["skip"]["rows"]
                base_col = ac + rule["skip"]["cols"]

                start_row = base_row + rule["extract"]["start"]["rows"]
                start_col = base_col + rule["extract"]["start"]["cols"]
                stop_col = base_col + rule["extract"]["stop_col"]

                rows = extract_block(sheet, start_row, start_col, stop_col)

                block = []
                for row in rows:
                    row_dict = {f"col_{i+1}": v for i, v in enumerate(row)}
                    block.append(row_dict)

                extracted[rule["as"]].append(block)

        with open("extracted_output.json", "w") as f:
            json.dump(extracted, f, indent=4)

        messagebox.showinfo("Done", "extracted_output.json created")


# ================== RUN ==================

if __name__ == "__main__":
    root = tk.Tk()
    app = MappingWizardApp(root)
    root.mainloop()