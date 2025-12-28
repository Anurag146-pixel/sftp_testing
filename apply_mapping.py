import json
from openpyxl import load_workbook

EXCEL_FILE = "EMB_improvedexample_2025.xlsx"
RULES_FILE = "mapping_rules.json"
OUTPUT_FILE = "extracted_output.json"


def load_excel_rows(file_path):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    return rows


def find_keyword_row(rows, keyword):
    for i, row in enumerate(rows):
        joined = " ".join(row)
        if keyword in joined:
            return i
    return -1


def extract_until(rows, start_row, skip_rows, stop_keyword):
    data = []
    i = start_row + skip_rows + 1
    while i < len(rows):
        joined = " ".join(rows[i])
        if stop_keyword and stop_keyword in joined:
            break
        if any(cell.strip() for cell in rows[i]):
            data.append(rows[i])
        i += 1
    return data


def main():
    rows = load_excel_rows(EXCEL_FILE)

    with open(RULES_FILE, "r") as f:
        rules = json.load(f)

    output = {}

    for rule in rules:
        keyword = rule["keyword"]
        rule_type = rule["rule_type"]
        skip_rows = rule.get("skip_rows", 0)
        stop_before = rule.get("stop_before", "")
        section = rule["section"]

        start_row = find_keyword_row(rows, keyword)

        if start_row == -1:
            output[section] = []
            continue

        if rule_type == "extract_until":
            output[section] = extract_until(
                rows,
                start_row,
                skip_rows,
                stop_before
            )

    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print("✅ Excel data extracted successfully")
    print("📄 Output file:", OUTPUT_FILE)


if __name__ == "__main__":
    main()